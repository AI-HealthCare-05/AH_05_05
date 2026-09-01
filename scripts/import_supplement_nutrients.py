from __future__ import annotations

import argparse
import asyncio
import sys
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Literal

from numbers_parser import Document
from openpyxl import load_workbook
from tortoise import Tortoise
from tortoise.backends.base.client import BaseDBAsyncClient
from tortoise.transactions import in_transaction

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.core.db.databases import TORTOISE_ORM  # noqa: E402
from app.models.supplement_nutrients import SupplementNutrient  # noqa: E402

EXPECTED_HEADERS = [
    "식품코드",
    "식품명",
    "영양성분제공단위량",
    "에너지(kcal)",
    "수분(g)",
    "단백질(g)",
    "지방(g)",
    "회분(g)",
    "탄수화물(g)",
    "당류(g)",
    "식이섬유(g)",
    "칼슘(mg)",
    "철(mg)",
    "인(mg)",
    "칼륨(mg)",
    "나트륨(mg)",
    "비타민 A(μg RAE)",
    "레티놀(μg)",
    "베타카로틴(μg)",
    "티아민(mg)",
    "리보플라빈(mg)",
    "니아신(mg)",
    "비타민 C(mg)",
    "비타민 D(μg)",
    "콜레스테롤(mg)",
    "포화지방산(g)",
    "트랜스지방산(g)",
    "1회분량",
    "1회분량중량/부피",
    "1일섭취횟수",
    "섭취대상",
]


@dataclass(frozen=True)
class FieldSpec:
    name: str
    kind: Literal["string", "integer", "decimal"]
    required: bool
    max_length: int | None = None
    max_digits: int | None = None
    decimal_places: int | None = None


FIELD_SPECS = [
    FieldSpec("food_code", "string", True, max_length=20),
    FieldSpec("name", "string", True, max_length=100),
    FieldSpec("basis_qty", "string", True, max_length=10),
    FieldSpec("energy_kcal", "integer", True),
    FieldSpec("water_g", "decimal", False, max_digits=10, decimal_places=3),
    FieldSpec("protein_g", "decimal", True, max_digits=5, decimal_places=2),
    FieldSpec("fat_g", "decimal", False, max_digits=5, decimal_places=2),
    FieldSpec("ash_g", "decimal", False, max_digits=10, decimal_places=3),
    FieldSpec("carb_g", "decimal", True, max_digits=6, decimal_places=2),
    FieldSpec("sugar_g", "decimal", False, max_digits=5, decimal_places=2),
    FieldSpec("fiber_g", "decimal", False, max_digits=7, decimal_places=1),
    FieldSpec("calcium_mg", "integer", False),
    FieldSpec("iron_mg", "decimal", False, max_digits=5, decimal_places=2),
    FieldSpec("phosphorus_mg", "integer", False),
    FieldSpec("potassium_mg", "integer", False),
    FieldSpec("sodium_mg", "integer", False),
    FieldSpec("vitamin_a_ug_rae", "integer", False),
    FieldSpec("retinol_ug", "integer", False),
    FieldSpec("beta_carotene_ug", "integer", False),
    FieldSpec("thiamine_mg", "decimal", False, max_digits=6, decimal_places=3),
    FieldSpec("riboflavin_mg", "decimal", False, max_digits=6, decimal_places=3),
    FieldSpec("niacin_mg", "decimal", False, max_digits=6, decimal_places=3),
    FieldSpec("vitamin_c_mg", "decimal", False, max_digits=7, decimal_places=2),
    FieldSpec("vitamin_d_ug", "decimal", False, max_digits=7, decimal_places=2),
    FieldSpec("cholesterol_mg", "decimal", False, max_digits=6, decimal_places=2),
    FieldSpec("sat_fat_g", "decimal", False, max_digits=4, decimal_places=2),
    FieldSpec("trans_fat_g", "decimal", False, max_digits=4, decimal_places=2),
    FieldSpec("serving_desc", "string", True, max_length=10),
    FieldSpec("serving_size", "string", True, max_length=10),
    FieldSpec("daily_freq", "string", True, max_length=5),
    FieldSpec("target", "string", False, max_length=10),
]

SOURCE_FIELDS = [spec.name for spec in FIELD_SPECS]
MFDS_HEADER_BY_FIELD = {
    "food_code": "식품코드",
    "name": "식품명",
    "basis_qty": "영양성분제공단위량",
    "energy_kcal": "에너지(kcal)",
    "water_g": "수분(g)",
    "protein_g": "단백질(g)",
    "fat_g": "지방(g)",
    "ash_g": "회분(g)",
    "carb_g": "탄수화물(g)",
    "sugar_g": "당류(g)",
    "fiber_g": "식이섬유(g)",
    "calcium_mg": "칼슘(mg)",
    "iron_mg": "철(mg)",
    "phosphorus_mg": "인(mg)",
    "potassium_mg": "칼륨(mg)",
    "sodium_mg": "나트륨(mg)",
    "vitamin_a_ug_rae": "비타민A(μg RAE)",
    "retinol_ug": "레티놀(μg)",
    "beta_carotene_ug": "베타카로틴(μg)",
    "thiamine_mg": "티아민(mg)",
    "riboflavin_mg": "리보플라빈(mg)",
    "niacin_mg": "니아신(mg)",
    "vitamin_c_mg": "비타민 C(mg)",
    "vitamin_d_ug": "비타민 D(μg)",
    "cholesterol_mg": "콜레스테롤(mg)",
    "sat_fat_g": "포화지방산(g)",
    "trans_fat_g": "트랜스지방산(g)",
    # 최신 식약처 파일에는 제형 단위(1정/1캡슐)가 따로 없으므로
    # 원문 1회분량중량/부피를 그대로 보존하고 제형을 추측하지 않는다.
    "serving_desc": "1회분량중량/부피",
    "serving_size": "1회분량중량/부피",
    "daily_freq": "1일섭취횟수",
    "target": "섭취대상명",
}


class ImportValidationError(ValueError):
    pass


@dataclass(frozen=True)
class ImportResult:
    total: int
    created: int
    updated: int
    failed: int = 0


def validate_headers(headers: list[object], *, row_number: int) -> None:
    normalized = [str(value).strip() if value is not None else None for value in headers]
    if normalized != EXPECTED_HEADERS:
        raise ImportValidationError(f"header mismatch at row {row_number}")


def _decimal_digit_counts(value: Decimal) -> tuple[int, int]:
    sign, digits, exponent = value.as_tuple()
    del sign
    decimal_places = max(-exponent, 0)
    integer_places = max(len(digits) + exponent, 0)
    return max(integer_places + decimal_places, 1), decimal_places


def _convert_value(value: object, spec: FieldSpec, *, row_number: int, header: str) -> object:
    if value is None or (isinstance(value, str) and not value.strip()):
        if spec.required:
            raise ImportValidationError(f"row {row_number}, column {header}: value is required")
        return None

    if spec.kind == "string":
        normalized = str(value).strip()
        if spec.max_length is not None and len(normalized) > spec.max_length:
            raise ImportValidationError(
                f"row {row_number}, column {header}: length {len(normalized)} exceeds {spec.max_length}"
            )
        return normalized

    try:
        number = Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise ImportValidationError(f"row {row_number}, column {header}: invalid number {value!r}") from exc
    if not number.is_finite():
        raise ImportValidationError(f"row {row_number}, column {header}: number must be finite")

    if spec.kind == "integer":
        if number != number.to_integral_value():
            raise ImportValidationError(f"row {row_number}, column {header}: integer required")
        return int(number)

    total_digits, decimal_places = _decimal_digit_counts(number)
    if total_digits > spec.max_digits or decimal_places > spec.decimal_places:
        raise ImportValidationError(
            f"row {row_number}, column {header}: {number} exceeds decimal({spec.max_digits},{spec.decimal_places})"
        )
    return number


def validate_rows(rows: list[list[object]], *, first_row_number: int) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    code_rows: dict[str, int] = {}
    for index, row in enumerate(rows):
        row_number = first_row_number + index
        if not any(value is not None and (not isinstance(value, str) or value.strip()) for value in row):
            continue
        if len(row) != len(FIELD_SPECS):
            raise ImportValidationError(f"row {row_number}: expected {len(FIELD_SPECS)} columns, received {len(row)}")
        record = {
            spec.name: _convert_value(value, spec, row_number=row_number, header=header)
            for value, spec, header in zip(row, FIELD_SPECS, EXPECTED_HEADERS, strict=True)
        }
        food_code = str(record["food_code"])
        if food_code in code_rows:
            raise ImportValidationError(
                f"duplicate food code {food_code} at rows {code_rows[food_code]} and {row_number}"
            )
        code_rows[food_code] = row_number
        records.append(record)
    return records


def validate_mfds_rows(
    headers: list[object],
    rows: list[list[object]],
    *,
    header_row_number: int,
) -> list[dict[str, object]]:
    normalized_headers = [str(value).strip() if value is not None else "" for value in headers]
    header_indexes = {header: index for index, header in enumerate(normalized_headers)}
    required_headers = set(MFDS_HEADER_BY_FIELD.values())
    missing_headers = sorted(required_headers.difference(header_indexes))
    if missing_headers:
        raise ImportValidationError(f"header mismatch at row {header_row_number}: missing={','.join(missing_headers)}")

    selected_rows = [
        [
            row[header_indexes[MFDS_HEADER_BY_FIELD[spec.name]]]
            if header_indexes[MFDS_HEADER_BY_FIELD[spec.name]] < len(row)
            else None
            for spec in FIELD_SPECS
        ]
        for row in rows
    ]
    return validate_rows(selected_rows, first_row_number=header_row_number + 1)


def _first_nonempty_row_index(rows: list[list[object]]) -> int:
    header_index = next(
        (index for index, row in enumerate(rows) if any(value is not None and str(value).strip() for value in row)),
        None,
    )
    if header_index is None:
        raise ImportValidationError("header row not found")
    return header_index


def read_numbers_rows(path: str | Path) -> list[list[object]]:
    document = Document(str(path))
    return document.sheets[0].tables[0].rows(values_only=True)


def read_xlsx_rows(path: str | Path) -> list[list[object]]:
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        return [list(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def parse_workbook(path: str | Path) -> list[dict[str, object]]:
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix == ".numbers":
        rows = read_numbers_rows(source_path)
    elif suffix == ".xlsx":
        rows = read_xlsx_rows(source_path)
    else:
        raise ImportValidationError(f"unsupported workbook type: {suffix or '<none>'}")

    header_index = _first_nonempty_row_index(rows)
    headers = rows[header_index]
    normalized_headers = [str(value).strip() if value is not None else None for value in headers]
    if normalized_headers == EXPECTED_HEADERS:
        return validate_rows(rows[header_index + 1 :], first_row_number=header_index + 2)
    return validate_mfds_rows(
        headers,
        rows[header_index + 1 :],
        header_row_number=header_index + 1,
    )


def parse_numbers(path: str | Path) -> list[dict[str, object]]:
    return parse_workbook(path)


async def upsert_records(records: list[dict[str, object]], *, batch_size: int = 500) -> ImportResult:
    codes = [str(record["food_code"]) for record in records]
    async with in_transaction() as connection:
        existing = await SupplementNutrient.filter(food_code__in=codes).using_db(connection)
        existing_by_code = {product.food_code: product for product in existing}
        to_create: list[SupplementNutrient] = []
        to_update: list[SupplementNutrient] = []
        for record in records:
            product = existing_by_code.get(str(record["food_code"]))
            if product is None:
                to_create.append(SupplementNutrient(**record))
                continue
            for field_name in SOURCE_FIELDS:
                setattr(product, field_name, record[field_name])
            to_update.append(product)
        if to_create:
            await SupplementNutrient.bulk_create(to_create, batch_size=batch_size, using_db=connection)
        if to_update:
            await SupplementNutrient.bulk_update(
                to_update,
                fields=SOURCE_FIELDS,
                batch_size=batch_size,
                using_db=connection,
            )
        await verify_stored_records(records, connection=connection)
    return ImportResult(total=len(records), created=len(to_create), updated=len(to_update))


async def verify_stored_records(
    records: list[dict[str, object]],
    *,
    connection: BaseDBAsyncClient | None = None,
) -> None:
    expected_codes = {str(record["food_code"]) for record in records}
    query = SupplementNutrient.filter(food_code__in=expected_codes)
    if connection is not None:
        query = query.using_db(connection)
    stored_codes = set(await query.values_list("food_code", flat=True))
    missing_codes = expected_codes.difference(stored_codes)
    if missing_codes:
        raise RuntimeError(f"건강기능식품 원본 식품코드가 DB에 모두 저장되지 않았습니다: missing={len(missing_codes)}")


def _validate_expected_count(records: list[dict[str, object]], expected_count: int | None) -> None:
    if expected_count is not None and len(records) != expected_count:
        raise ImportValidationError(f"expected {expected_count} records, received {len(records)}")


async def _run_import(
    path: Path,
    *,
    dry_run: bool,
    expected_count: int | None,
) -> tuple[list[dict[str, object]], ImportResult | None]:
    records = parse_workbook(path)
    _validate_expected_count(records, expected_count)
    if dry_run:
        return records, None
    await Tortoise.init(config=TORTOISE_ORM)
    try:
        return records, await upsert_records(records)
    finally:
        await Tortoise.close_connections()


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="건강기능식품 영양성분 Numbers/XLSX 파일을 MySQL에 적재합니다.")
    parser.add_argument("path", type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--expected-count", type=int)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    try:
        records, result = asyncio.run(
            _run_import(
                args.path,
                dry_run=args.dry_run,
                expected_count=args.expected_count,
            )
        )
    except Exception as exc:
        print("total=0 created=0 updated=0 failed=1", file=sys.stdout)
        print(str(exc), file=sys.stderr)
        return 1
    if result is None:
        print(f"total={len(records)} dry_run=true", file=sys.stdout)
        return 0
    print(
        f"total={result.total} created={result.created} updated={result.updated} failed={result.failed}",
        file=sys.stdout,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
